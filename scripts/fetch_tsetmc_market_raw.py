"""Fetch raw TSETMC snapshots for the full market universe.

Flow:
  1) Download TSETMC MarketWatchPlus (all instruments).
  2) Use each row's insCode as the primary identity.
  3) Verify the instrument identity with TSETMC InstrumentInfo.
  4) Use the verified TSETMC symbol as the folder name.
  5) Store the complete raw response as a daily snapshot:
       runtime/بورس_خام/<verified_symbol>/<YYYY-MM>/<YYYY-MM-DD>.json

The raw payload is not normalized or analyzed.
"""
from __future__ import annotations

import datetime as dt
import json
import re
import time
from pathlib import Path
from urllib.parse import quote
from urllib.request import Request, urlopen

BASE = "https://cdn.tsetmc.com/api"
MARKETWATCH_URL = "https://old.tsetmc.com/tsev2/excel/MarketWatchPlus.aspx?d=0"
HEADERS = {"User-Agent": "Mozilla/5.0 SMART-full-market-raw/1.0"}
ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "runtime" / "بورس_خام"
UNIVERSE_DIR = ROOT / "runtime" / "بورس_خام" / "_universe"

ENDPOINTS = {
    "instrument_info": "/Instrument/GetInstrumentInfo/{ins_code}",
    "closing_info": "/ClosingPrice/GetClosingPriceInfo/{ins_code}",
    "daily_history": "/ClosingPrice/GetClosingPriceDailyList/{ins_code}/60",
    "client_type": "/ClientType/GetClientType/{ins_code}/1/0",
}

RETRY_STATUS = {429, 500, 502, 503, 504}


def get_bytes(url: str, attempts: int = 4) -> bytes:
    last_exc: Exception | None = None
    for attempt in range(attempts):
        try:
            req = Request(url, headers=HEADERS)
            with urlopen(req, timeout=45) as response:
                return response.read()
        except Exception as exc:
            last_exc = exc
            if attempt + 1 < attempts:
                time.sleep(1.5 * (2 ** attempt))
    raise last_exc or RuntimeError("request failed")


def get_json(path: str) -> object:
    data = get_bytes(BASE + path)
    return json.loads(data.decode("utf-8"))


def gregorian_to_jalali(gy: int, gm: int, gd: int) -> tuple[int, int, int]:
    g_days = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334]
    gy2 = gy + 1 if gm > 2 else gy
    days = 355666 + 365 * gy + (gy2 + 3) // 4 - (gy2 + 99) // 100 + (gy2 + 399) // 400 + gd + g_days[gm - 1]
    jy = -1595 + 33 * (days // 12053)
    days %= 12053
    jy += 4 * (days // 1461)
    days %= 1461
    if days > 365:
        jy += (days - 1) // 365
        days = (days - 1) % 365
    if days < 186:
        jm = 1 + days // 31
        jd = 1 + days % 31
    else:
        jm = 7 + (days - 186) // 30
        jd = 1 + (days - 186) % 30
    return jy, jm, jd


def norm(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).replace("\u200c", "")


def parse_marketwatch(raw: bytes) -> list[dict[str, str]]:
    """Parse the MarketWatchPlus xlsx without pandas."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required: pip install openpyxl") from exc

    import io
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = None
    header_idx = None
    wanted = {
        "inscode": {"inscode", "ins_code", "کدسهام", "کدسهم", "کد"},
        "symbol": {"نماد", "symbol", "lval18afc", "lval18"},
    }
    for i, row in enumerate(rows[:15]):
        normalized = {norm(v).lower(): j for j, v in enumerate(row) if v is not None}
        if any(k in normalized for k in wanted["inscode"]) and any(k in normalized for k in wanted["symbol"]):
            header_row = normalized
            header_idx = i
            break

    if header_row is None:
        # Fallback: known MarketWatchPlus layout has insCode and lVal18 columns.
        for i, row in enumerate(rows[:10]):
            vals = [norm(v).lower() for v in row]
            if any("inscode" in v or "کد" in v for v in vals) and any("نماد" in v or "lval18" in v for v in vals):
                header_idx = i
                header_row = {v: j for j, v in enumerate(vals)}
                break
    if header_row is None:
        raise RuntimeError("Could not identify MarketWatchPlus columns")

    def find_col(keys: set[str]) -> int:
        for key, idx in header_row.items():
            if key in keys or any(k in key for k in keys if len(k) >= 4):
                return idx
        raise RuntimeError(f"Required MarketWatch column not found: {keys}")

    ins_idx = find_col(wanted["inscode"])
    sym_idx = find_col(wanted["symbol"])

    out: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in rows[header_idx + 1:]:
        if ins_idx >= len(row) or sym_idx >= len(row):
            continue
        ins = norm(row[ins_idx])
        sym = str(row[sym_idx] or "").strip()
        if not ins.isdigit() or not sym or ins in seen:
            continue
        seen.add(ins)
        out.append({"ins_code": ins, "market_symbol": sym})
    return out


def verified_instrument(ins_code: str, market_symbol: str) -> tuple[str, object]:
    payload = get_json(ENDPOINTS["instrument_info"].format(ins_code=ins_code))
    info = payload.get("instrumentInfo", {}) if isinstance(payload, dict) else {}
    verified_symbol = str(info.get("lVal18AFC") or "").strip()
    returned_code = str(info.get("insCode") or ins_code)
    if returned_code != ins_code:
        raise RuntimeError(f"insCode mismatch: market={ins_code}, info={returned_code}")
    if not verified_symbol:
        raise RuntimeError(f"TSETMC returned no symbol for insCode={ins_code}")
    if norm(verified_symbol) != norm(market_symbol):
        # MarketWatch can contain formatting differences; the InstrumentInfo symbol is authoritative.
        print(f"VERIFY: {market_symbol} -> {verified_symbol} ({ins_code})")
    return verified_symbol, payload


def main() -> int:
    now = dt.datetime.now(dt.timezone.utc)
    jy, jm, jd = gregorian_to_jalali(now.year, now.month, now.day)
    date_str = f"{jy:04d}-{jm:02d}-{jd:02d}"
    month_dir = f"{jy:04d}-{jm:02d}"

    print("Downloading full TSETMC market universe...")
    universe_raw = get_bytes(MARKETWATCH_URL)
    instruments = parse_marketwatch(universe_raw)
    if not instruments:
        raise SystemExit("No instruments found in MarketWatchPlus")

    UNIVERSE_DIR.mkdir(parents=True, exist_ok=True)
    (UNIVERSE_DIR / f"{date_str}.json").write_text(
        json.dumps({"source": "TSETMC MarketWatchPlus", "received_at_utc": now.isoformat(), "count": len(instruments), "instruments": instruments}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Universe: {len(instruments)} instruments")

    ok = 0
    failed = 0
    for n, item in enumerate(instruments, 1):
        ins_code = item["ins_code"]
        market_symbol = item["market_symbol"]
        try:
            symbol, instrument_info = verified_instrument(ins_code, market_symbol)
            payload = {
                "source": "TSETMC",
                "symbol": symbol,
                "ins_code": ins_code,
                "market_symbol": market_symbol,
                "market_date_jalali": date_str,
                "received_at_utc": now.isoformat(),
                "raw": {
                    "market_universe_row": item,
                    "instrument_info": instrument_info,
                    "closing_info": get_json(ENDPOINTS["closing_info"].format(ins_code=ins_code)),
                    "daily_history": get_json(ENDPOINTS["daily_history"].format(ins_code=ins_code)),
                    "client_type": get_json(ENDPOINTS["client_type"].format(ins_code=ins_code)),
                },
            }
            out_dir = RAW / symbol / month_dir
            out_dir.mkdir(parents=True, exist_ok=True)
            out_file = out_dir / f"{date_str}.json"
            out_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            ok += 1
            print(f"[{n}/{len(instruments)}] OK: {symbol} -> {out_file}")
        except Exception as exc:
            failed += 1
            print(f"[{n}/{len(instruments)}] ERROR: {market_symbol} ({ins_code}): {exc}")

    print(f"Finished: ok={ok}, failed={failed}, total={len(instruments)}")
    return 1 if ok == 0 else 0


if __name__ == "__main__":
    raise SystemExit(main())
