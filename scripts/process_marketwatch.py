"""Process the saved raw TSETMC MarketWatch file into verified symbol snapshots.

Input:
  runtime/market_raw/marketwatch/YYYY-MM-DD.gz

Output:
  runtime/market_raw/stocks/<verified_symbol>/<YYYY-MM>/<YYYY-MM-DD>.json
  runtime/monitored/<verified_symbol>/latest.json

Identity rule:
  insCode is the primary key. The folder name is the verified TSETMC symbol
  returned by InstrumentInfo. Therefore symbols such as فولاد and فولادح cannot
  be merged merely because their display names are similar.

The original MarketWatch .gz remains untouched.
"""
from __future__ import annotations

import datetime as dt
import gzip
import io
import json
import re
from pathlib import Path
from urllib.parse import quote
from urllib.request import Request, urlopen

BASE = "https://cdn.tsetmc.com/api"
HEADERS = {"User-Agent": "Mozilla/5.0 SMART-marketwatch-processor/1.0"}
ROOT = Path(__file__).resolve().parents[1]
MARKETWATCH_DIR = ROOT / "runtime" / "market_raw" / "marketwatch"
STOCKS_DIR = ROOT / "runtime" / "market_raw" / "stocks"
MONITORED_DIR = ROOT / "runtime" / "monitored"
UNIVERSE_DIR = ROOT / "runtime" / "market_raw" / "universe"

ENDPOINTS = {
    "instrument_info": "/Instrument/GetInstrumentInfo/{ins_code}",
    "closing_info": "/ClosingPrice/GetClosingPriceInfo/{ins_code}",
    "daily_history": "/ClosingPrice/GetClosingPriceDailyList/{ins_code}/60",
    "client_type": "/ClientType/GetClientType/{ins_code}/1/0",
}


def get_json(path: str) -> object:
    req = Request(BASE + path, headers=HEADERS)
    with urlopen(req, timeout=45) as response:
        return json.loads(response.read().decode("utf-8"))


def norm(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).replace("\u200c", "")


def gregorian_to_jalali(gy: int, gm: int, gd: int) -> tuple[int, int, int]:
    g_days = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334]
    gy2 = gy + 1 if gm > 2 else gy
    days = 355666 + 365 * gy + (gy2 + 3) // 4 - (gy2 + 99) // 100 + (gy2 + 399) // 400 + gd + g_days[gm - 1]
    jy = -1595 + 33 * (days // 12053)
    days %= 12053
    jy += 4 * (days // 1461)
    days %= 1461
    if days > 365:
        jy += (days - 1) // 365
        days = (days - 1) % 365
    if days < 186:
        jm = 1 + days // 31
        jd = 1 + days % 31
    else:
        jm = 7 + (days - 186) // 30
        jd = 1 + (days - 186) % 30
    return jy, jm, jd


def parse_marketwatch(raw: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = None
    header_idx = None
    for i, row in enumerate(rows[:20]):
        vals = [norm(v).lower() for v in row]
        has_code = any(v in {"inscode", "ins_code", "کدسهام", "کدسهم"} or "inscode" in v for v in vals)
        has_symbol = any(v in {"نماد", "symbol", "lval18afc", "lval18"} or "lval18" in v for v in vals)
        if has_code and has_symbol:
            header_idx = i
            header_row = {v: j for j, v in enumerate(vals) if v}
            break
    if header_row is None:
        raise RuntimeError("Could not identify MarketWatchPlus columns")

    def find_col(kind: str) -> int:
        keys = ("inscode", "ins_code", "کدسهام", "کدسهم") if kind == "code" else ("نماد", "symbol", "lval18afc", "lval18")
        for key, idx in header_row.items():
            if key in keys or any(k in key for k in keys if len(k) >= 4):
                return idx
        raise RuntimeError(f"Required MarketWatch column not found: {kind}")

    ins_idx = find_col("code")
    sym_idx = find_col("symbol")
    out: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in rows[header_idx + 1:]:
        if max(ins_idx, sym_idx) >= len(row):
            continue
        ins = norm(row[ins_idx])
        sym = str(row[sym_idx] or "").strip()
        if not ins.isdigit() or not sym or ins in seen:
            continue
        seen.add(ins)
        out.append({"ins_code": ins, "market_symbol": sym})
    return out


def verify(ins_code: str, market_symbol: str) -> tuple[str, object]:
    payload = get_json(ENDPOINTS["instrument_info"].format(ins_code=quote(ins_code)))
    info = payload.get("instrumentInfo", {}) if isinstance(payload, dict) else {}
    returned_code = str(info.get("insCode") or ins_code)
    symbol = str(info.get("lVal18AFC") or "").strip()
    if returned_code != ins_code:
        raise RuntimeError(f"insCode mismatch: market={ins_code}, info={returned_code}")
    if not symbol:
        raise RuntimeError(f"No verified symbol for insCode={ins_code}")
    return symbol, payload


def main() -> int:
    files = sorted(MARKETWATCH_DIR.glob("*.gz"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise SystemExit(f"No raw MarketWatch file found in {MARKETWATCH_DIR}")
    source = files[0]
    print(f"Processing: {source}")

    raw = gzip.decompress(source.read_bytes())
    instruments = parse_marketwatch(raw)
    if not instruments:
        raise SystemExit("No instruments found in raw MarketWatch")

    stamp = dt.datetime.now(dt.timezone.utc)
    jy, jm, jd = gregorian_to_jalali(stamp.year, stamp.month, stamp.day)
    date_str = f"{jy:04d}-{jm:02d}-{jd:02d}"
    month_dir = f"{jy:04d}-{jm:02d}"

    UNIVERSE_DIR.mkdir(parents=True, exist_ok=True)
    (UNIVERSE_DIR / f"{date_str}.json").write_text(
        json.dumps({"source_file": str(source.relative_to(ROOT)), "count": len(instruments), "instruments": instruments}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Universe: {len(instruments)} instruments")

    ok = failed = 0
    for n, item in enumerate(instruments, 1):
        ins_code = item["ins_code"]
        try:
            symbol, instrument_info = verify(ins_code, item["market_symbol"])
            payload = {
                "source": "TSETMC",
                "symbol": symbol,
                "ins_code": ins_code,
                "market_symbol": item["market_symbol"],
                "market_date_jalali": date_str,
                "source_file": str(source.relative_to(ROOT)),
                "received_at_utc": stamp.isoformat(),
                "raw": {
                    "market_universe_row": item,
                    "instrument_info": instrument_info,
                    "closing_info": get_json(ENDPOINTS["closing_info"].format(ins_code=quote(ins_code))),
                    "daily_history": get_json(ENDPOINTS["daily_history"].format(ins_code=quote(ins_code))),
                    "client_type": get_json(ENDPOINTS["client_type"].format(ins_code=quote(ins_code))),
                },
            }
            out = STOCKS_DIR / symbol / month_dir / f"{date_str}.json"
            out.parent.mkdir(parents=True, exist_ok=True)
            out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

            monitored = MONITORED_DIR / symbol
            monitored.mkdir(parents=True, exist_ok=True)
            (monitored / "latest.json").write_text(json.dumps({
                "symbol": symbol,
                "ins_code": ins_code,
                "source_file": str(out.relative_to(ROOT)),
                "updated_at_utc": stamp.isoformat(),
            }, ensure_ascii=False, indent=2), encoding="utf-8")
            ok += 1
            print(f"[{n}/{len(instruments)}] OK {symbol} ({ins_code})")
        except Exception as exc:
            failed += 1
            print(f"[{n}/{len(instruments)}] ERROR {item['market_symbol']} ({ins_code}): {exc}")

    print(f"Finished: ok={ok}, failed={failed}, total={len(instruments)}")
    return 1 if ok == 0 else 0


if __name__ == "__main__":
    raise SystemExit(main())
